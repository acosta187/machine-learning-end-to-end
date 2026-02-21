"""
SCRIPT 1: Generación del Glosario de Jerga Peruana
===================================================
Genera glosario.csv y glosario.xlsx con términos, definiciones,
polaridad sentimental y ejemplos de uso en contexto.

Uso: python 01_generar_glosario.py
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================
# DATOS DEL GLOSARIO
# ==============================================================
# Estructura: (termino, equivalencia_estandar, polaridad, categoria, ejemplos...)
# polaridad: POSITIVE, NEGATIVE, NEUTRAL, AMBIGUOUS

GLOSARIO = [
    # ------------------------------------------------------------------ POSITIVOS
    {
        "termino": "bacán",
        "equivalencia": "excelente, genial, increíble",
        "polaridad": "POSITIVE",
        "categoria": "valoración general",
        "ejemplo_1": "Ese concierto estuvo bacán, la pasé genial.",
        "ejemplo_2": "Qué bacán que te hayan dado el trabajo.",
        "ejemplo_3": "Tu amigo es bien bacán, siempre te apoya.",
        "notas": "Muy frecuente en habla juvenil limeña. Intensificado: 'bacanazo'."
    },
    {
        "termino": "chévere",
        "equivalencia": "bueno, agradable, cool",
        "polaridad": "POSITIVE",
        "categoria": "valoración general",
        "ejemplo_1": "El restaurante está chévere, la comida rica.",
        "ejemplo_2": "Qué chévere que hayas venido.",
        "ejemplo_3": "Es un chico chévere, no te va a fallar.",
        "notas": "Compartido con Venezuela y Colombia. Muy natural en Perú."
    },
    {
        "termino": "mostro",
        "equivalencia": "excelente, muy hábil, impresionante",
        "polaridad": "POSITIVE",
        "categoria": "habilidad/desempeño",
        "ejemplo_1": "Eres un mostro para la programación.",
        "ejemplo_2": "Ese jugador es un mostro, mete gol siempre.",
        "ejemplo_3": "Quedó mostro el proyecto, felicitaciones.",
        "notas": "Forma coloquial de 'monstruo'. Muy usado en Lima."
    },
    {
        "termino": "fiera",
        "equivalencia": "persona muy hábil o destacada",
        "polaridad": "POSITIVE",
        "categoria": "habilidad/desempeño",
        "ejemplo_1": "Es una fiera en matemáticas.",
        "ejemplo_2": "Esa empresa es una fiera en el mercado.",
        "ejemplo_3": "Mi profe de cálculo es una fiera, todo explica clarito.",
        "notas": "Puede usarse para personas o entidades. Siempre positivo."
    },
    {
        "termino": "piola",
        "equivalencia": "tranquilo, discreto, sin problemas, confiable",
        "polaridad": "POSITIVE",
        "categoria": "actitud/comportamiento",
        "ejemplo_1": "El evento estuvo piola, sin dramas.",
        "ejemplo_2": "Es un pata piola, no molesta a nadie.",
        "ejemplo_3": "Todo piola por acá, no te preocupes.",
        "notas": "Puede ser NEUTRAL en contextos descriptivos. Muy versátil."
    },
    {
        "termino": "de pelos",
        "equivalencia": "muy bien, excelente, perfecto",
        "polaridad": "POSITIVE",
        "categoria": "valoración general",
        "ejemplo_1": "Me cayó de pelos ese favor que me hiciste.",
        "ejemplo_2": "La reunión salió de pelos.",
        "ejemplo_3": "El diseño quedó de pelos, bien profesional.",
        "notas": "Expresión muy peruana. Sinónimo coloquial de 'de lujo'."
    },
    {
        "termino": "cachaciento",
        "equivalencia": "gracioso, chistoso, con buen humor",
        "polaridad": "POSITIVE",
        "categoria": "personalidad",
        "ejemplo_1": "Ese pata es bien cachaciento, siempre te saca una risa.",
        "ejemplo_2": "La reunión estuvo cachaciento, nadie se aburrió.",
        "ejemplo_3": "Me cae bien, es cachaciento sin ser pesado.",
        "notas": "Exclusivamente peruano. No confundir con burlón."
    },
    {
        "termino": "arriba",
        "equivalencia": "de primer nivel, excelente calidad",
        "polaridad": "POSITIVE",
        "categoria": "valoración general",
        "ejemplo_1": "Ese profesor es arriba, explica clarito.",
        "ejemplo_2": "El servicio estuvo arriba, sin quejas.",
        "ejemplo_3": "La atención en ese hotel es arriba.",
        "notas": "Menos frecuente, uso más adulto. Indica calidad superior."
    },
    {
        "termino": "causa",
        "equivalencia": "amigo, pata, compañero",
        "polaridad": "POSITIVE",
        "categoria": "relaciones sociales",
        "ejemplo_1": "Mi causa siempre está cuando lo necesito.",
        "ejemplo_2": "Qué causa eres, gracias por ayudarme.",
        "ejemplo_3": "Ese es mi causa del colegio, lo conozco de años.",
        "notas": "Término de afecto. Indica cercanía. El contexto define la polaridad."
    },
    {
        "termino": "pata",
        "equivalencia": "amigo, colega, persona",
        "polaridad": "POSITIVE",
        "categoria": "relaciones sociales",
        "ejemplo_1": "Mi pata me prestó su laptop para el examen.",
        "ejemplo_2": "Es un buen pata, siempre te da la mano.",
        "ejemplo_3": "¿Quién es ese pata? No lo conozco.",
        "notas": "Muy neutro si solo identifica a alguien. Positivo cuando hay afecto implícito."
    },
    {
        "termino": "pulento",
        "equivalencia": "excelente, de buena calidad, impecable",
        "polaridad": "POSITIVE",
        "categoria": "valoración general",
        "ejemplo_1": "El show estuvo pulento, todos quedaron contentos.",
        "ejemplo_2": "Ese ceviche está pulento, mejor que en el centro.",
        "ejemplo_3": "Tu presentación quedó pulenta, bien hecha.",
        "notas": "Usado más en generaciones jóvenes. Intensificado: 'pulentucho'."
    },
    {
        "termino": "alzado",
        "equivalencia": "que subió de nivel, que mejoró su posición",
        "polaridad": "POSITIVE",
        "categoria": "logro/éxito",
        "ejemplo_1": "Ya estás alzado con ese cargo nuevo.",
        "ejemplo_2": "Desde que le fue bien en el negocio está alzado.",
        "ejemplo_3": "Mi hermano ya está alzado, tiene su casa propia.",
        "notas": "POSITIVO en contexto de logro. NEGATIVO si implica soberbia ('se ha alzado')."
    },

    # ------------------------------------------------------------------ NEGATIVOS
    {
        "termino": "palta",
        "equivalencia": "vergüenza, incomodidad, pena ajena",
        "polaridad": "NEGATIVE",
        "categoria": "emoción negativa",
        "ejemplo_1": "Me dio palta verlo tropezar frente a todos.",
        "ejemplo_2": "Qué palta que hayas llegado tarde a tu propia reunión.",
        "ejemplo_3": "Eso me da palta, no lo hagas en público.",
        "notas": "Exclusivamente peruano. Puede ser leve (vergüenza ajena) o intensa."
    },
    {
        "termino": "roche",
        "equivalencia": "vergüenza, situación incómoda, bochorno",
        "polaridad": "NEGATIVE",
        "categoria": "emoción negativa",
        "ejemplo_1": "Qué roche que te hayan visto así.",
        "ejemplo_2": "Me dio un roche enorme cuando me llamaron la atención.",
        "ejemplo_3": "Eso fue un roche total, todos se quedaron mirando.",
        "notas": "Muy similar a 'palta'. 'Roche' tiende a ser más social/público."
    },
    {
        "termino": "maleado",
        "equivalencia": "en mal estado, descompuesto, de mala calidad",
        "polaridad": "NEGATIVE",
        "categoria": "calidad/estado",
        "ejemplo_1": "El producto llegó maleado, no sirve.",
        "ejemplo_2": "Ese local está maleado, no vayas.",
        "ejemplo_3": "La comida estaba maleada, me cayó mal.",
        "notas": "Se aplica a objetos, lugares y personas. Para personas: mal educado, corrompido."
    },
    {
        "termino": "misio",
        "equivalencia": "sin dinero, en mala situación económica",
        "polaridad": "NEGATIVE",
        "categoria": "situación económica",
        "ejemplo_1": "Estoy bien misio este mes, no puedo salir.",
        "ejemplo_2": "No puedo comprarlo, estoy misio.",
        "ejemplo_3": "Siempre está misio pero igual gasta.",
        "notas": "Describe estado temporal. Neutro en tono descriptivo, negativo en queja."
    },
    {
        "termino": "sapo",
        "equivalencia": "entrometido, chismoso, delator",
        "polaridad": "NEGATIVE",
        "categoria": "personalidad negativa",
        "ejemplo_1": "No le cuentes nada, es bien sapo.",
        "ejemplo_2": "El sapo ese fue y contó todo a la jefa.",
        "ejemplo_3": "Odio a los sapos en el trabajo.",
        "notas": "Siempre peyorativo en este uso. No confundir con el animal."
    },
    {
        "termino": "cargoso",
        "equivalencia": "molesto, insistente, pesado",
        "polaridad": "NEGATIVE",
        "categoria": "personalidad negativa",
        "ejemplo_1": "Ese cliente es bien cargoso, llama a cada rato.",
        "ejemplo_2": "No seas cargoso, ya te dije que no.",
        "ejemplo_3": "Su jefe es un cargoso, siempre está encima.",
        "notas": "Muy frecuente. Puede intensificarse: 'bien cargoso', 'súper cargoso'."
    },
    {
        "termino": "maleta",
        "equivalencia": "inútil, incapaz, que no rinde",
        "polaridad": "NEGATIVE",
        "categoria": "habilidad negativa",
        "ejemplo_1": "Ese técnico es un maleta, no arregló nada.",
        "ejemplo_2": "No lo pongas de titular, es un maleta.",
        "ejemplo_3": "Me salió maleta el proveedor, hay que cambiarlo.",
        "notas": "Se usa para personas principalmente. Fuerte carga despectiva."
    },
    {
        "termino": "feo",
        "equivalencia": "malo, grave, inaceptable (no físico)",
        "polaridad": "NEGATIVE",
        "categoria": "valoración negativa",
        "ejemplo_1": "Eso está feo, no debería pasar.",
        "ejemplo_2": "Qué feo te trató, eso no se hace.",
        "ejemplo_3": "El ambiente está feo en esa empresa.",
        "notas": "En este uso NO se refiere al aspecto físico. Indica situación grave o trato injusto."
    },
    {
        "termino": "pesado",
        "equivalencia": "fastidioso, molesto, insoportable",
        "polaridad": "NEGATIVE",
        "categoria": "personalidad negativa",
        "ejemplo_1": "Es un pesado, siempre quejándose de todo.",
        "ejemplo_2": "Qué pesado ese proceso de atención.",
        "ejemplo_3": "Me cae pesado ese gerente.",
        "notas": "Compartido con España y Latinoamérica. Muy común en Perú."
    },
    {
        "termino": "chancho",
        "equivalencia": "sucio, descuidado, de mala presentación",
        "polaridad": "NEGATIVE",
        "categoria": "calidad/higiene",
        "ejemplo_1": "Ese local está chancho, no vuelvo.",
        "ejemplo_2": "Come bien chancho, no tiene modales.",
        "ejemplo_3": "Qué chancho ese baño, dan ganas de salir corriendo.",
        "notas": "En uso no literal. 'Chancho' como adjetivo = sucio/descuidado."
    },
    {
        "termino": "de mala",
        "equivalencia": "con mala suerte, situación desafortunada",
        "polaridad": "NEGATIVE",
        "categoria": "suerte/circunstancia",
        "ejemplo_1": "De mala que llovió justo hoy.",
        "ejemplo_2": "Me fue de mala en el examen.",
        "ejemplo_3": "Siempre le va de mala en los negocios.",
        "notas": "Puede ser NEUTRAL si solo describe mala suerte sin queja emocional."
    },
    {
        "termino": "tronado",
        "equivalencia": "loco, fuera de sí, irracional",
        "polaridad": "NEGATIVE",
        "categoria": "estado mental",
        "ejemplo_1": "Ese tipo está tronado, no le hagas caso.",
        "ejemplo_2": "Qué propuesta más tronada, nadie la va a aprobar.",
        "ejemplo_3": "Se tronó con la presión del trabajo.",
        "notas": "Puede ser despectivo o descriptivo. Contexto define intensidad."
    },
    {
        "termino": "dar cólera",
        "equivalencia": "dar rabia, indignar, molestar profundamente",
        "polaridad": "NEGATIVE",
        "categoria": "emoción negativa",
        "ejemplo_1": "Me da cólera que siempre lleguen tarde.",
        "ejemplo_2": "Esa noticia me dio cólera todo el día.",
        "ejemplo_3": "Le da cólera que no lo tomen en cuenta.",
        "notas": "'Cólera' como sustantivo = rabia intensa. Muy peruano."
    },
    {
        "termino": "estar en llanta",
        "equivalencia": "estar sin dinero, en quiebra",
        "polaridad": "NEGATIVE",
        "categoria": "situación económica",
        "ejemplo_1": "Este mes estoy en llanta, no me alcanza nada.",
        "ejemplo_2": "Después de las fiestas quedé en llanta.",
        "ejemplo_3": "La empresa está en llanta, van a cerrar.",
        "notas": "Similar a 'misio' pero con mayor intensidad. Más formal en uso."
    },

    # ------------------------------------------------------------------ NEUTRALES / AMBIGUOS
    {
        "termino": "tela",
        "equivalencia": "mucho, bastante, en gran cantidad",
        "polaridad": "NEUTRAL",
        "categoria": "intensificador",
        "ejemplo_1": "Hay tela de gente en el centro hoy.",
        "ejemplo_2": "Tiene tela de trabajo acumulado.",
        "ejemplo_3": "Costó tela de plata ese viaje.",
        "notas": "Intensificador neutro. La polaridad la da el contexto del sustantivo."
    },
    {
        "termino": "causa",
        "equivalencia": "amigo, pata (forma de llamar a alguien)",
        "polaridad": "NEUTRAL",
        "categoria": "relaciones sociales",
        "ejemplo_1": "Oye causa, ¿tienes un rato?",
        "ejemplo_2": "¿Qué tal causa? ¿Cómo te fue?",
        "ejemplo_3": "Causa, te llamo luego.",
        "notas": "Como vocativo es NEUTRAL. Con valoración explícita puede ser POSITIVE."
    },
    {
        "termino": "pata",
        "equivalencia": "persona, tipo (sin connotación)",
        "polaridad": "NEUTRAL",
        "categoria": "referencia a persona",
        "ejemplo_1": "¿Quién es ese pata del fondo?",
        "ejemplo_2": "Un pata me dijo que mañana no hay clase.",
        "ejemplo_3": "No sé qué pata hizo eso.",
        "notas": "NEUTRAL cuando solo identifica a alguien sin valoración afectiva."
    },
    {
        "termino": "al toque",
        "equivalencia": "de inmediato, enseguida, rápido",
        "polaridad": "NEUTRAL",
        "categoria": "tiempo/velocidad",
        "ejemplo_1": "Me llamó al toque que llegó.",
        "ejemplo_2": "Lo resolvieron al toque, sin demora.",
        "ejemplo_3": "Responde al toque los mensajes.",
        "notas": "Puramente adverbial. Positivo si el contexto valora la rapidez."
    },
    {
        "termino": "pues",
        "equivalencia": "partícula enfática / marcador discursivo",
        "polaridad": "NEUTRAL",
        "categoria": "marcador discursivo",
        "ejemplo_1": "Ya pues, si tú lo dices.",
        "ejemplo_2": "¿Qué pasó pues?",
        "ejemplo_3": "Anda pues, no seas así.",
        "notas": "Marcador discursivo muy peruano. No aporta polaridad, pero marca registro informal."
    },
    {
        "termino": "pe",
        "equivalencia": "pues (forma reducida, enfática)",
        "polaridad": "NEUTRAL",
        "categoria": "marcador discursivo",
        "ejemplo_1": "Ya pe, qué quieres que haga.",
        "ejemplo_2": "¿Qué fue pe?",
        "ejemplo_3": "Hazlo pe, no te cuesta nada.",
        "notas": "Apócope de 'pues'. Marca informalidad y registro oral."
    },
    {
        "termino": "ahorita",
        "equivalencia": "en un momento (tiempo impreciso)",
        "polaridad": "NEUTRAL",
        "categoria": "tiempo",
        "ejemplo_1": "Ahorita te atiendo.",
        "ejemplo_2": "Ahorita llega, espera.",
        "ejemplo_3": "Ahorita lo hago, dame un segundo.",
        "notas": "Tiempo indefinido. Puede implicar demora. Neutro en sí mismo."
    },
    {
        "termino": "hablar al chancho",
        "equivalencia": "hablar mal de alguien, criticar a sus espaldas",
        "polaridad": "NEGATIVE",
        "categoria": "comportamiento social negativo",
        "ejemplo_1": "Le hablan al chancho a su jefe todo el día.",
        "ejemplo_2": "No me gusta la gente que habla al chancho.",
        "ejemplo_3": "Se enteró que le hablaban al chancho en la oficina.",
        "notas": "Expresión idiomática. Siempre negativa por el acto que describe."
    },
    {
        "termino": "estar en su salsa",
        "equivalencia": "estar cómodo, disfrutar de algo",
        "polaridad": "POSITIVE",
        "categoria": "estado emocional",
        "ejemplo_1": "En esa reunión estaba en su salsa, hablando de lo que le gusta.",
        "ejemplo_2": "Cocinando está en su salsa, no lo pares.",
        "ejemplo_3": "Cuando habla de fútbol está en su salsa.",
        "notas": "Compartido con español estándar pero muy usado en Perú."
    },
    {
        "termino": "hacer la vista gorda",
        "equivalencia": "ignorar algo a propósito, tolerar algo incorrecto",
        "polaridad": "NEGATIVE",
        "categoria": "comportamiento social",
        "ejemplo_1": "El supervisor hace la vista gorda con las tardanzas.",
        "ejemplo_2": "No podemos seguir haciendo la vista gorda con eso.",
        "ejemplo_3": "Siempre le hacen la vista gorda por ser amigo del jefe.",
        "notas": "Negativo por implicación de complicidad o falta de integridad."
    },
]

# ==============================================================
# GENERAR CSV
# ==============================================================
import os  # 👈 añade esto arriba junto a los imports

def generar_csv(glosario, ruta="sintetic_data/glosario_jerga_peruana.csv"):
    directorio = os.path.dirname(ruta)
    if directorio:
        os.makedirs(directorio, exist_ok=True)

    df = pd.DataFrame(glosario)

    cols = ["termino", "equivalencia", "polaridad", "categoria",
            "ejemplo_1", "ejemplo_2", "ejemplo_3", "notas"]
    df = df[cols]

    df.to_csv(ruta, index=False, encoding="utf-8-sig")
    print(f"✅ CSV generado: {ruta} ({len(df)} términos)")
    return df


# ==============================================================
# GENERAR XLSX CON FORMATO
# ==============================================================
COLOR_POSITIVE = "C6EFCE"   # verde suave
COLOR_NEGATIVE = "FFC7CE"   # rojo suave
COLOR_NEUTRAL  = "FFEB9C"   # amarillo suave
COLOR_HEADER   = "2E4057"   # azul oscuro
COLOR_HEADER_FONT = "FFFFFF"

def estilo_header(cell):
    cell.font = Font(bold=True, color=COLOR_HEADER_FONT, name="Arial", size=11)
    cell.fill = PatternFill("solid", start_color=COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def estilo_fila(cell, polaridad):
    color = {"POSITIVE": COLOR_POSITIVE, "NEGATIVE": COLOR_NEGATIVE}.get(polaridad, COLOR_NEUTRAL)
    cell.fill = PatternFill("solid", start_color=color)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.font = Font(name="Arial", size=10)

def borde_delgado():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def generar_xlsx(df, ruta="glosario_jerga_peruana.xlsx"):
    wb = Workbook()

    # ---- HOJA 1: Glosario completo ----
    ws = wb.active
    ws.title = "Glosario Completo"

    headers = ["Término", "Equivalencia Estándar", "Polaridad", "Categoría",
               "Ejemplo 1", "Ejemplo 2", "Ejemplo 3", "Notas"]
    col_widths = [16, 30, 12, 22, 45, 45, 45, 40]

    for i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=i, value=h)
        estilo_header(cell)
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for row_idx, row in df.iterrows():
        excel_row = row_idx + 2
        ws.row_dimensions[excel_row].height = 55
        valores = [row["termino"], row["equivalencia"], row["polaridad"],
                   row["categoria"], row["ejemplo_1"], row["ejemplo_2"],
                   row["ejemplo_3"], row["notas"]]
        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=valor)
            estilo_fila(cell, row["polaridad"])
            cell.border = borde_delgado()
            if col_idx == 3:  # columna Polaridad: negrita
                cell.font = Font(bold=True, name="Arial", size=10)

    # ---- HOJA 2: Solo términos + polaridad (para tokenizer) ----
    ws2 = wb.create_sheet("Para Tokenizer")
    ws2["A1"] = "Término"
    ws2["B1"] = "Polaridad"
    ws2["C1"] = "Agregar al Vocab"
    for cell in [ws2["A1"], ws2["B1"], ws2["C1"]]:
        estilo_header(cell)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 18

    vistos = set()
    fila = 2
    for _, row in df.iterrows():
        t = row["termino"]
        if t not in vistos:
            ws2.cell(row=fila, column=1, value=t)
            ws2.cell(row=fila, column=2, value=row["polaridad"])
            ws2.cell(row=fila, column=3, value="Sí")
            for col in range(1, 4):
                estilo_fila(ws2.cell(row=fila, column=col), row["polaridad"])
                ws2.cell(row=fila, column=col).border = borde_delgado()
            vistos.add(t)
            fila += 1

    # ---- HOJA 3: Resumen estadístico ----
    ws3 = wb.create_sheet("Resumen")
    ws3["A1"] = "Resumen del Glosario"
    ws3["A1"].font = Font(bold=True, size=14, color=COLOR_HEADER, name="Arial")
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 14

    stats = df["polaridad"].value_counts().to_dict()
    ws3["A3"] = "Polaridad"
    ws3["B3"] = "Cantidad"
    for c in [ws3["A3"], ws3["B3"]]:
        estilo_header(c)

    fila = 4
    for pol, cnt in stats.items():
        ws3.cell(row=fila, column=1, value=pol)
        ws3.cell(row=fila, column=2, value=cnt)
        for col in range(1, 3):
            estilo_fila(ws3.cell(row=fila, column=col), pol)
        fila += 1

    ws3.cell(row=fila, column=1, value="TOTAL")
    ws3.cell(row=fila, column=2, value=f"=SUM(B4:B{fila-1})")
    ws3.cell(row=fila, column=1).font = Font(bold=True, name="Arial")
    ws3.cell(row=fila, column=2).font = Font(bold=True, name="Arial")

    ws3["A8"] = "Categorías únicas"
    ws3["B8"] = df["categoria"].nunique()
    ws3["A9"] = "Términos únicos"
    ws3["B9"] = df["termino"].nunique()

    wb.save(ruta)
    print(f"✅ XLSX generado: {ruta} (3 hojas)")

# ==============================================================
# MAIN
# ==============================================================
if __name__ == "__main__":
    ruta_csv = "sintetic_data/glosario_jerga_peruana.csv"
    ruta_xlsx = "sintetic_data/glosario_jerga_peruana.xlsx"

    df = generar_csv(GLOSARIO, ruta_csv)
    generar_xlsx(df, ruta_xlsx)

    print("\n📊 Distribución por polaridad:")
    print(df["polaridad"].value_counts().to_string())
    print(f"\n📋 Total de entradas: {len(df)}")

